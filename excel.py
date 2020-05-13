from openpyxl import load_workbook
import os
import Tkinter, tkFileDialog

def getList(ws):
    result = "BUY: "
    for i in range(2, 17):
        cell1 = 'B' + str(i)
        cell2 = 'D' + str(i)
        # print(cell)
        name = str(ws[cell2].value).strip()
        ticker = ws[cell1].value
        result += name + " " + ticker
        if i < 16:
            result += " | "
    result += "\n\nSELL: "
    for i in range(2, 17):
        cell1 = 'I' + str(i)
        cell2 = 'K' + str(i)
        # print(cell)
        name = ws[cell2].value
        name = str(name)
        name = name.strip()
        ticker = ws[cell1].value
        result += name + " " + ticker
        if i < 16:
            result += " | "
    return result

def main():
    root = Tkinter.Tk()
    root.withdraw()

    my_file = tkFileDialog.askopenfilename()

    # # filename = raw_input('Input the excel file name: ')
    # THIS_FOLDER = os.path.dirname(os.path.abspath(__file__))
    # # print(THIS_FOLDER)
    # my_file = os.path.join(THIS_FOLDER, 'test.xlsm')
    # print(my_file)

    wb = load_workbook(my_file)
    # print(wb.sheetnames)
    FINI = wb.sheetnames[0]
    ITC = wb.sheetnames[1]
    ws1 = wb.get_sheet_by_name(FINI)
    ws2 = wb.get_sheet_by_name(ITC)
    # print(ws1['B2'].value)
    print("Taiwan Fund Flow Today")
    result = getList(ws1)
    print("FINI TOP 15 BUY/SELL by Value Today\n" + result + "\n")
    result = getList(ws2)
    print("ITC TOP 15 BUY/SELL by Value Today\n" + result)


main()